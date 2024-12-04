import json

from django.shortcuts import render, redirect
from django.http import JsonResponse
from openpyxl.styles import Alignment, Font, Border, Side

from menu.models import Position, Category
from openpyxl import Workbook


def order_page(request):
    order = request.session.get('order', {})
    order = {int(k): int(v) for k, v in order.items()}
    user_info = request.session.get('user_info', {})
    order_list = Position.objects.filter(id__in=order.keys())

    context = {
        "position_list": Position.objects.all(),
        "pos_by_cat": Category.objects.prefetch_related('positions'),
        "category_list": Category.objects.all(),
        "order": order,
        "order_list": order_list,
        "user_info": user_info
    }
    return render(request, 'order.html', context=context)


def generate_xlsx(request):
    print("hi")
    wb = Workbook()
    ws = wb.active
    ws.title = 'Меню'
    order = request.session.get('order', {})
    order = {int(k): int(v) for k, v in order.items()}
    user_info = request.session.get('user_info', {})
    order_list = Position.objects.filter(id__in=order.keys())

    fn = 'Меню(' + user_info['user_name'] + '_' + user_info['date_of_event'] + ').xlsx'
    print(fn)

    font = Font(bold=True)
    for row in ws['A1': 'E3']:
        for cell in row:
            cell.font = font

    ws.column_dimensions['A'].width = 70
    ws['A1'] = 'Выездное банкетное меню'

    ws['C1'] = user_info['user_name']
    ws['D1'] = 'Дата мероприятия:'
    ws['D2'] = user_info['date_of_event']
    ws['E1'] = 'Количество гостей:'
    ws['E2'] = user_info['guests_quantity']
    ws.column_dimensions['B'].width = 17
    ws.column_dimensions['C'].width = 17
    ws.column_dimensions['D'].width = 17
    ws.column_dimensions['E'].width = 17
    ws['D1'].alignment = Alignment(horizontal='left', vertical='top')
    ws['E1'].alignment = Alignment(horizontal='left', vertical='top')

    for row in ws['A3':'D3']:
        for cell in row:
            cell = ' '

    ws.append(['Название', 'Порция', 'Количество', 'Цена', 'Итог'])
    for row in ws['A4': 'E4']:
        for cell in row:
            cell.font = font
            cell.border = Border(left=Side(border_style='medium', color='FF000000'),
            right=Side(border_style='medium', color='FF000000'),
            top=Side(border_style='medium', color='FF000000'),
            bottom=Side(border_style='medium', color='FF000000'))

    for i in order:
        for p in order_list:
            if i == p.id:
                ws.append([p.title, p.portion, order[i], p.price, p.price * order[i]])

    total_cost = 0
    for i in order:
        position = Position.objects.get(id=i)
        total_cost += position.price * order[i]
    ws.append(['', '', '', 'Итого:', total_cost])

    wb.save(fn)
    wb.close()
    return JsonResponse({'success': True})


def order_list_ajax(request):
    order = request.session.get('order', {})
    user_info = request.session.get('user_info', {})
    total_cost = 0
    for i in order:
        position = Position.objects.get(id=i)
        total_cost += position.price * order[i]

    return JsonResponse({'order': order, 'total_cost': total_cost, 'user_info': user_info})


def set_quantity_by_guests(request):
    order = request.session.get('order', {})
    user_info = request.session.get('user_info', {})
    for k in order:
        order[k] = user_info['guests_quantity']
    request.session['order'] = order
    request.session.modified = True

    return JsonResponse({'success': True})


# Добавить или увеличить количества товара на 1
def add_to_order(request, position_id):
    print("def add_to_order")
    if request.method == 'POST' and request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest':
        if 'order' not in request.session:
            request.session['order'] = {}
        order = request.session['order']
        position_id = str(position_id)
        if position_id in order:
            order[position_id] += 1
            print("in")
        else:
            order[position_id] = 1
            print("not in")

        request.session['order'] = order
        request.session.modified = True
        return JsonResponse({'success': True})
    else:
        return JsonResponse({'success': False})


# Уменьшить количество позиции на 1 или удалить ее из заказа
def remove_from_order(request, position_id):
    if request.method == 'POST' and request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest':
        order = request.session.get('order', {})
        position_id = str(position_id)
        last = False
        if position_id in order:
            order[position_id] -= 1
            if order[position_id] <= 0:
                del order[position_id]
                last = True
        request.session['order'] = order
        request.session.modified = True
        return JsonResponse({'success': True, 'last': last})
    else:
        return JsonResponse({'success': False})


def set_guests_quantity(request, guests_quantity):
    print("def set_guests_quantity")
    if request.method == 'POST' and request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest':
        user_info = request.session.get('user_info', {})
        user_info['guests_quantity'] = guests_quantity
        request.session['user_info'] = user_info
        request.session.modified = True
        print(user_info)
        return JsonResponse({'success': True})
    else:
        return JsonResponse({'success': False})


def set_user_name(request, user_name):
    print("def set_user_name")
    if request.method == 'POST' and request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest':
        user_info = request.session.get('user_info', {})
        user_info['user_name'] = user_name
        request.session['user_info'] = user_info
        request.session.modified = True
        print(user_info)
        return JsonResponse({'success': True})
    else:
        return JsonResponse({'success': False})


def set_date_of_event(request, date_of_event):
    print("def set_date_of_event")
    if request.method == 'POST' and request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest':
        user_info = request.session.get('user_info', {})
        user_info['date_of_event'] = date_of_event
        request.session['user_info'] = user_info
        request.session.modified = True
        print(user_info)
        return JsonResponse({'success': True})
    else:
        return JsonResponse({'success': False})


def set_quantity(request, position_id):
    print("def set quantity by guests quantity")
    if request.method == 'POST' and request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest':
        user_info = request.session.get('user_info', {})
        order = request.session['order']
        order[position_id] = user_info['guests_quantity']
        request.session['order'] = order
        request.session.modified = True
        return JsonResponse({'success': True})
    else:
        return JsonResponse({'success': False})


def update_quantity(request, position_id, quantity):
    print("def quantity updated")
    if request.method == 'POST' and request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest':
        order = request.session['order']
        order[position_id] = quantity
        request.session['order'] = order
        request.session.modified = True
        return JsonResponse({'success': True})
    else:
        return JsonResponse({'success': False})

    # if request.method == 'POST':
    #     if not request.session.get('order'):
    #         request.session['order'] = list()
    #     else:
    #         request.session['order'] = list(request.session['order'])
    #
    # # check if item exist in list of dicts
    #     item_exist = next((item for item in request.session['order'] if item["type"] == request.POST.get('type') and
    #                        id == item["id"]), False)
    #
    #     add_data = {
    #         'type': request.POST.get('type'),
    #         'id': id,
    #     }
    # # if id not in order_ids_list:
    #     if not item_exist:
    #         request.session['order'].append(add_data)
    #         request.session.modified = True
    #
    # # for AJAX requests
    # if request.is_ajax():
    #     data = {
    #         'type': request.POST.get('type'),
    #         'id': request.POST.get('id'),
    #     }
    #     request.session.modified = True
    #     return JsonResponse(data)
    # END for AJAX requests

# def remove_from_order(request, id):
#     if request.method == 'POST':
#         # Delete an item from order
#         for item in request.session['order']:
#             if item['id'] == id and item['type'] == request.POST.get('type'):
#                 item.clear()
#
#         # Remove empty {} from order_list
#         while {} in request.session['order']:
#             request.session['order'].remove({})
#
#         # Remove order if order_list is empty
#         if not request.session['order']:
#             del request.session['order']
#             # request.session.modified = True
#
#         request.session.modified = True
#
#         # for AJAX requests
#     if request.is_ajax():
#         data = {
#             'type': request.POST.get('type'),
#             'id': request.POST.get('id'),
#         }
#         request.session.modified = True
#         return JsonResponse(data)
#     # END for AJAX requests
#     return redirect(request.POST.get('url_from'))
#
#
def delete_order(request):
    if request.session.get('order'):
        del request.session['order']
        request.session.modified = True
    return redirect(request.POST.get('url_from'))


def order_api(request):
    return JsonResponse(request.session.get('order'), safe=False)
